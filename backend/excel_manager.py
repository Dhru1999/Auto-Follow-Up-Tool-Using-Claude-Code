"""
Excel Manager - Handles reading/writing job application data to/from Excel.
Watches the file for changes so data stays in sync without manual uploads.
"""

import os
import time
import threading
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUMNS = [
    "ID", "Company", "Role", "Applied Date", "Recruiter Name",
    "Recruiter Email", "HR Email", "Job URL", "Status",
    "Follow Up Date", "Follow Up Count", "Last Follow Up Date",
    "Notes", "AI Draft", "Email Sent"
]

STATUS_OPTIONS = [
    "Applied", "Follow Up Sent", "Interview Scheduled",
    "Interview Done", "Offer Received", "Rejected", "Withdrawn", "No Response"
]

COLUMN_ALIASES = {
    "company name": "Company",
    "companyname": "Company",
    "role": "Role",
    "role ": "Role",
    "job title": "Role",
    "date of applied": "Applied Date",
    "date applied": "Applied Date",
    "applied date": "Applied Date",
    "repsonse received": "Status",
    "response received": "Status",
    "feedback": "Status",
    "additional comments": "Notes",
    "comments": "Notes",
    "email sent": "Email Sent",
    "recruiter email": "Recruiter Email",
    "hr email": "HR Email",
    "follow up date": "Follow Up Date",
    "followup date": "Follow Up Date",
    "follow up count": "Follow Up Count",
    "last follow up date": "Last Follow Up Date",
    "application id": "ID",
    "id": "ID",
    "unnamed: 0": "ID",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _normalize_header(header: str) -> str:
    return str(header or "").strip().lower().replace("\u00A0", " ").replace("\r", " ").replace("\n", " ").strip()


def _map_header_name(header: str) -> str:
    normalized = _normalize_header(header)
    if normalized in COLUMN_ALIASES:
        return COLUMN_ALIASES[normalized]
    for standard in COLUMNS:
        if normalized == standard.strip().lower():
            return standard
    return str(header).strip()


def _map_sheet_columns(columns):
    return [_map_header_name(c) for c in columns]


def create_default_excel(file_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    ws.append(COLUMNS)
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    widths = {
        "ID": 6, "Company": 20, "Role": 25, "Applied Date": 14,
        "Recruiter Name": 20, "Recruiter Email": 28, "HR Email": 28,
        "Job URL": 35, "Status": 20, "Follow Up Date": 16,
        "Follow Up Count": 14, "Last Follow Up Date": 20,
        "Notes": 35, "AI Draft": 50, "Email Sent": 12
    }
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Add sample row
    today = date.today().strftime("%Y-%m-%d")
    ws.append([
        1, "Example Corp", "Software Engineer", today,
        "Jane Smith", "jane@example.com", "hr@example.com",
        "https://example.com/job/123", "Applied",
        "", 0, "", "Sample application — delete or edit this row", "", "No"
    ])

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["Auto Follow-Up Tool — Instructions"],
        [""],
        ["COLUMNS EXPLAINED:"],
        ["ID", "Unique number for each application (auto-assigned by tool)"],
        ["Company", "Company name"],
        ["Role", "Job title you applied for"],
        ["Applied Date", "Date you applied (YYYY-MM-DD)"],
        ["Recruiter Name", "Name of recruiter or hiring manager"],
        ["Recruiter Email", "Recruiter email (tool sends follow-up here)"],
        ["HR Email", "HR email (CC'd in follow-up)"],
        ["Job URL", "Link to the job posting"],
        ["Status", f"One of: {', '.join(STATUS_OPTIONS)}"],
        ["Follow Up Date", "Next scheduled follow-up date (auto-set)"],
        ["Follow Up Count", "How many follow-ups sent (auto-tracked)"],
        ["Last Follow Up Date", "Date of last follow-up (auto-tracked)"],
        ["Notes", "Your personal notes"],
        ["AI Draft", "AI-generated email draft (auto-filled by tool)"],
        ["Email Sent", "Yes/No — updated automatically after sending"],
        [""],
        ["TIPS:"],
        ["• Just add/edit rows in 'Job Applications' sheet and save — the tool picks up changes automatically"],
        ["• Set Status to 'Withdrawn' or 'Offer Received' to stop follow-ups"],
        ["• The tool checks every hour and sends follow-ups on the scheduled date"],
        ["• Gmail users: use App Passwords (not your regular password)"],
    ]
    for row in instructions:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    print(f"Created Excel file: {file_path}")


def load_applications(file_path: str) -> list[dict]:
    if not Path(file_path).exists():
        create_default_excel(file_path)

    # Read the sheet and preserve its actual columns so the frontend can
    # display exactly what exists in the Excel file (no extra or missing columns).
    df = pd.read_excel(file_path, sheet_name="Job Applications", dtype=str)
    df = df.fillna("")

    mapped_columns = _map_sheet_columns(df.columns.tolist())
    df.columns = mapped_columns
    sheet_cols = [str(c) for c in df.columns.tolist()]

    applications = []
    for _, row in df.iterrows():
        app = {col: str(row.get(col, "")).strip() for col in sheet_cols}
        if not (app.get("Company") or app.get("Role")):
            continue
        applications.append(app)

    return applications


def sync_source_status(source_path: str, target_path: str) -> int:
    """Sync non-empty Status values from a source Excel file into the target workbook."""
    src_path = Path(source_path)
    dst_path = Path(target_path)
    if not src_path.exists() or not dst_path.exists():
        return 0

    src_df = pd.read_excel(src_path, sheet_name=0, dtype=str).fillna("")
    src_df.columns = _map_sheet_columns(src_df.columns.tolist())
    if "Status" not in src_df.columns:
        return 0

    def build_key(row):
        if row.get("ID"):
            return (str(row.get("ID")).strip(), "")
        return (
            str(row.get("Company", "")).strip().lower(),
            str(row.get("Role", "")).strip().lower(),
            str(row.get("Applied Date", "")).strip().lower(),
        )

    src_lookup = {}
    for _, row in src_df.iterrows():
        key = build_key(row)
        status = str(row.get("Status", "")).strip()
        if status:
            src_lookup[key] = status

    wb = load_workbook(dst_path)
    ws = wb["Job Applications"]
    header = {
        _map_header_name(cell.value): cell.column
        for cell in ws[1] if cell.value
    }
    status_idx = header.get("Status")
    id_idx = header.get("ID")
    company_idx = header.get("Company")
    role_idx = header.get("Role")
    applied_idx = header.get("Applied Date")
    if not status_idx:
        return 0

    updated = 0
    for row in ws.iter_rows(min_row=2):
        key = None
        if id_idx:
            key = (str(row[id_idx - 1].value or "").strip(), "")
        if not key or not key[0]:
            key = (
                str(row[company_idx - 1].value or "").strip().lower() if company_idx else "",
                str(row[role_idx - 1].value or "").strip().lower() if role_idx else "",
                str(row[applied_idx - 1].value or "").strip().lower() if applied_idx else "",
            )

        status = src_lookup.get(key)
        if status:
            cell = row[status_idx - 1]
            if str(cell.value or "").strip() != status:
                cell.value = status
                updated += 1

    if updated > 0:
        wb.save(dst_path)
    return updated


def save_application_update(file_path: str, app_id: str, updates: dict) -> bool:
    try:
        wb = load_workbook(file_path)
        ws = wb["Job Applications"]

        header = {
            _map_header_name(cell.value): cell.column
            for cell in ws[1] if cell.value
        }
        id_col = header.get("ID")
        if not id_col:
            print("Error saving update: ID column not found")
            return False

        for row in ws.iter_rows(min_row=2):
            row_id = str(row[id_col - 1].value or "").strip()
            if row_id == str(app_id):
                for field, value in updates.items():
                    col_idx = header.get(field)
                    if col_idx:
                        row[col_idx - 1].value = value
                break

        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error saving update: {e}")
        return False


def add_application(file_path: str, app_data: dict) -> int:
    if not Path(file_path).exists():
        create_default_excel(file_path)

    wb = load_workbook(file_path)
    ws = wb["Job Applications"]

    # Get next ID
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if row[0]:
                max_id = max(max_id, int(row[0]))
        except (ValueError, TypeError):
            pass
    new_id = max_id + 1

    header = [_map_header_name(cell.value) for cell in ws[1] if cell.value]
    row_values = []
    for canonical in header:
        if canonical == "ID":
            row_values.append(new_id)
        elif canonical == "Email Sent":
            row_values.append(app_data.get(canonical, "No") or "No")
        elif canonical == "Follow Up Count":
            row_values.append(app_data.get(canonical, 0) or 0)
        else:
            row_values.append(app_data.get(canonical, ""))

    # If the sheet header doesn't include all expected columns, append the remaining defaults.
    for col in COLUMNS:
        if col not in header:
            if col == "ID":
                row_values.append(new_id)
            elif col == "Email Sent":
                row_values.append(app_data.get(col, "No") or "No")
            elif col == "Follow Up Count":
                row_values.append(app_data.get(col, 0) or 0)
            else:
                row_values.append(app_data.get(col, ""))

    row_num = ws.max_row + 1
    ws.append(row_values)
    if row_num % 2 == 0:
        for cell in ws[row_num]:
            cell.fill = ALT_ROW_FILL

    wb.save(file_path)
    return new_id


class ExcelWatcher:
    """Watches the Excel file for changes and triggers a callback."""

    def __init__(self, file_path: str, callback, poll_interval: int = 30):
        self.file_path = file_path
        self.callback = callback
        self.poll_interval = poll_interval
        self._last_mtime = 0
        self._thread = None
        self._running = False

    def start(self):
        self._running = True
        self._thread = threading.Thread(target=self._watch, daemon=True)
        self._thread.start()

    def stop(self):
        self._running = False

    def _watch(self):
        while self._running:
            try:
                mtime = os.path.getmtime(self.file_path)
                if mtime != self._last_mtime:
                    if self._last_mtime != 0:
                        print(f"Excel file changed — reloading data...")
                        self.callback()
                    self._last_mtime = mtime
            except FileNotFoundError:
                pass
            time.sleep(self.poll_interval)
